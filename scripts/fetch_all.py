#!/usr/bin/env python3
"""Bermuda Market Intel — Data Pipeline v8.
Improved India / UK rates sourcing and roll-aware commodity futures history.
"""
import json, re, sys, os, logging, time, io, zipfile, csv
from datetime import datetime, timedelta, date
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed
import urllib.request

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
log = logging.getLogger("fetch")
DATA = Path(__file__).parent.parent / "data"
DATA.mkdir(exist_ok=True)
HDR = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
}

def get(url, timeout=8):
    req = urllib.request.Request(url, headers=HDR)
    with urllib.request.urlopen(req, timeout=timeout) as r:
        return r.read().decode("utf-8", errors="replace")

def get_bytes(url, timeout=20, retries=2, sleep_seconds=2):
    last_err = None
    for attempt in range(retries + 1):
        try:
            req = urllib.request.Request(url, headers=HDR)
            with urllib.request.urlopen(req, timeout=timeout) as r:
                return r.read()
        except Exception as e:
            last_err = e
            if attempt < retries:
                time.sleep(sleep_seconds * (attempt + 1))
    raise last_err

def write(name, obj):
    obj["_fetched"] = datetime.utcnow().isoformat() + "Z"
    (DATA / name).write_text(json.dumps(obj, indent=2, default=str))
    log.info(f"  wrote {name}")

def fred_csv(series_id, start="2024-01-01", retries=0):
    """Fetch single FRED series CSV."""
    for attempt in range(retries + 1):
        try:
            url = f"https://fred.stlouisfed.org/graph/fredgraph.csv?id={series_id}&cosd={start}"
            raw = get(url, timeout=6)
            obs = []
            for line in raw.strip().split("\n")[1:]:
                parts = line.split(",")
                if len(parts) >= 2 and parts[1] not in (".", ""):
                    try:
                        obs.append({"date": parts[0], "value": float(parts[1])})
                    except Exception:
                        pass
            obs.sort(key=lambda x: x["date"], reverse=True)
            if obs:
                return obs
        except Exception as e:
            log.warning(f"  FRED {series_id} attempt {attempt+1}: {e}")
            if attempt < retries:
                time.sleep(2)
    return []

def fred_multi_csv(series_ids, start="2024-01-01", retries=0):
    """Fetch MULTIPLE FRED series in ONE request. Returns {series_id: [obs]}."""
    joined = ",".join(series_ids)
    url = f"https://fred.stlouisfed.org/graph/fredgraph.csv?id={joined}&cosd={start}"
    result = {sid: [] for sid in series_ids}
    sid_upper = {sid.upper(): sid for sid in series_ids}
    for attempt in range(retries + 1):
        try:
            raw = get(url, timeout=8)
            rows = list(csv.reader(io.StringIO(raw)))
            if len(rows) < 2:
                return result
            header = [h.strip().strip('"') for h in rows[0]]
            col_map = {}
            for i, h in enumerate(header):
                if i == 0:
                    continue
                sid = sid_upper.get(h.upper())
                if sid:
                    col_map[i] = sid
            for parts in rows[1:]:
                if len(parts) < 2:
                    continue
                date_val = parts[0].strip().strip('"')
                if not re.match(r"\d{4}-\d{2}-\d{2}", date_val):
                    continue
                for ci, sid in col_map.items():
                    if ci < len(parts):
                        val = parts[ci].strip().strip('"')
                        if val not in (".", ""):
                            try:
                                result[sid].append({"date": date_val, "value": float(val)})
                            except Exception:
                                pass
            for sid in result:
                result[sid].sort(key=lambda x: x["date"], reverse=True)
            return result
        except Exception as e:
            log.warning(f"  FRED multi fetch attempt {attempt+1}: {e}")
            if attempt < retries:
                time.sleep(2 * (attempt + 1))
    return result

def fred_download_csv(series_id, start="2024-01-01"):
    """Alternative FRED endpoint: series download page instead of graph CSV.
    Served via a different URL path — useful when the graph endpoint is throttled.
    Single attempt, short timeout, no retries."""
    url = f"https://fred.stlouisfed.org/series/{series_id}/downloaddata/{series_id}.csv"
    try:
        raw = get(url, timeout=6)
        obs = []
        for line in raw.strip().split("\n")[1:]:
            parts = line.split(",")
            if len(parts) >= 2 and parts[1].strip() not in (".", ""):
                try:
                    d = parts[0].strip()
                    if d >= start:
                        obs.append({"date": d, "value": float(parts[1].strip())})
                except Exception:
                    pass
        obs.sort(key=lambda x: x["date"], reverse=True)
        if obs:
            log.info(f"  FRED↓ {series_id}: got {len(obs)} obs via download endpoint")
        return obs
    except Exception as e:
        log.warning(f"  FRED↓ {series_id}: {e}")
        return []

def fred_year_ago_10y(series_id):
    target = (datetime.utcnow() - timedelta(days=365)).strftime("%Y-%m-%d")
    obs = fred_csv(series_id, start=(datetime.utcnow() - timedelta(days=400)).strftime("%Y-%m-%d"))
    if not obs:
        return None, ""
    best = min(obs, key=lambda o: abs((datetime.strptime(o["date"], "%Y-%m-%d") - datetime.strptime(target, "%Y-%m-%d")).days))
    return round(best["value"], 4), best["date"]

def find_prior_date_yields(rows, days_ago, tenors, max_diff_days=14):
    """Find yields from rows list closest to N days ago. rows: [{date, yields}] sorted desc."""
    target = (datetime.utcnow() - timedelta(days=days_ago)).strftime("%Y-%m-%d")
    if not rows:
        return [None] * len(tenors), ""
    best = min(rows, key=lambda r: abs((datetime.strptime(r["date"], "%Y-%m-%d") - datetime.strptime(target, "%Y-%m-%d")).days))
    diff = abs((datetime.strptime(best["date"], "%Y-%m-%d") - datetime.strptime(target, "%Y-%m-%d")).days)
    if diff > max_diff_days:
        return [None] * len(tenors), ""
    return [best["yields"].get(t) for t in tenors], best["date"]

def fred_prior_single(series_id, days_ago, max_diff_days=14):
    """Fetch a single FRED series value closest to N days ago. Returns (value, date)."""
    target_dt = datetime.utcnow() - timedelta(days=days_ago)
    start = (target_dt - timedelta(days=30)).strftime("%Y-%m-%d")
    obs = fred_csv(series_id, start=start, retries=0)
    if not obs:
        return None, ""
    target = target_dt.strftime("%Y-%m-%d")
    best = min(obs, key=lambda o: abs((datetime.strptime(o["date"], "%Y-%m-%d") - datetime.strptime(target, "%Y-%m-%d")).days))
    diff = abs((datetime.strptime(best["date"], "%Y-%m-%d") - datetime.strptime(target, "%Y-%m-%d")).days)
    if diff > max_diff_days:
        return None, ""
    return round(best["value"], 4), best["date"]

def get_prior_spot(fred_series, yahoo_symbol, days_ago, max_diff_days=14):
    """FRED-first prior spot price; falls back to Yahoo Finance close."""
    v, d = fred_prior_single(fred_series, days_ago, max_diff_days=max_diff_days)
    if v is not None:
        return v, d
    return yahoo_price_at(yahoo_symbol, days_ago, max_diff_days=max_diff_days)

def validate_yield(val):
    """Range-check a yield value. Returns rounded float or None."""
    if val is None:
        return None
    return round(val, 4) if 0 < val < 20 else None

def interpolate_curve(yields_dict, tenors):
    """Linear interpolation for missing interior tenors. Does not extrapolate."""
    def t2n(t):
        return float(t.replace("Y", ""))
    known = sorted([(t2n(t), v) for t, v in yields_dict.items() if v is not None], key=lambda k: k[0])
    if len(known) < 2:
        return yields_dict
    for t in tenors:
        if yields_dict.get(t) is None:
            x = t2n(t)
            left = max((k for k in known if k[0] < x), default=None, key=lambda k: k[0])
            right = min((k for k in known if k[0] > x), default=None, key=lambda k: k[0])
            if left and right:
                yields_dict[t] = round(left[1] + (right[1] - left[1]) * (x - left[0]) / (right[0] - left[0]), 4)
    return yields_dict

def load_last_india():
    try:
        f = DATA / "india.json"
        if f.exists():
            return json.loads(f.read_text())
    except Exception:
        pass
    return None

def load_last_bma():
    try:
        f = DATA / "bma_rates.json"
        if f.exists():
            return json.loads(f.read_text())
    except Exception:
        pass
    return None

def load_last_gilt():
    try:
        f = DATA / "gilt.json"
        if f.exists():
            return json.loads(f.read_text())
    except Exception:
        pass
    return None

def load_last_commodities():
    try:
        f = DATA / "commodities.json"
        if f.exists():
            return json.loads(f.read_text())
    except Exception:
        pass
    return None

def load_last_credit():
    try:
        f = DATA / "credit.json"
        if f.exists():
            return json.loads(f.read_text())
    except Exception:
        pass
    return None

def append_curve_history(name, date_str, tenors, yields, source):
    path = DATA / f"{name}_history.jsonl"
    row = {
        "date": date_str,
        "source": source,
        "tenors": tenors,
        "yields": yields,
        "_fetched": datetime.utcnow().isoformat() + "Z",
    }
    try:
        existing = set()
        if path.exists():
            for line in path.read_text().splitlines():
                if line.strip():
                    obj = json.loads(line)
                    existing.add(obj.get("date"))
        if date_str not in existing:
            with path.open("a", encoding="utf-8") as f:
                f.write(json.dumps(row) + "\n")
    except Exception as e:
        log.warning(f"  history append failed for {name}: {e}")

def load_curve_history(name):
    path = DATA / f"{name}_history.jsonl"
    out = []
    try:
        if path.exists():
            for line in path.read_text().splitlines():
                if not line.strip():
                    continue
                obj = json.loads(line)
                out.append({
                    "date": obj["date"],
                    "yields": dict(zip(obj.get("tenors", []), obj.get("yields", [])))
                })
    except Exception as e:
        log.warning(f"  history load failed for {name}: {e}")
    out.sort(key=lambda x: x["date"], reverse=True)
    return out

def history_lookup(name, days_ago, tenors, max_diff_days=14):
    rows = load_curve_history(name)
    return find_prior_date_yields(rows, days_ago, tenors, max_diff_days=max_diff_days)

def parse_d(s):
    """Parse '31 December 2025' or '31 Dec 2025'. Returns datetime or None."""
    if not s:
        return None
    for fmt in ["%d %B %Y", "%d %b %Y"]:
        try:
            return datetime.strptime(s.strip(), fmt)
        except ValueError:
            pass
    return None

def extract_date_from_url(url):
    """Extract date from PDF filenames like Discount_Rates_31_December_2025.pdf."""
    m = re.search(r'(\d{1,2})[_\s-](\w+)[_\s-](\d{4})', url)
    if not m:
        return None
    return parse_d(f"{m.group(1)} {m.group(2)} {m.group(3)}")

def _to_iso_date(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%d-%b-%Y", "%d-%B-%Y", "%d/%m/%Y", "%m/%d/%Y", "%d.%m.%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except Exception:
            pass
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})", s)
    if m:
        return s[:10]
    return None

def _as_float(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "")
    s = s.replace("%", "")
    if s in ("", ".", "-", "--", "NA", "N/A"):
        return None
    try:
        return float(s)
    except Exception:
        return None

def _to_years(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        x = float(v)
        return x if 0 < x <= 100 else None
    s = str(v).strip().lower()
    s = s.replace("years", "y").replace("year", "y").replace("yrs", "y").replace("yr", "y")
    s = s.replace("months", "m").replace("month", "m").replace("mos", "m").replace("mo", "m")
    s = s.replace(" ", "")
    if re.fullmatch(r"\d+(\.\d+)?y", s):
        return float(s[:-1])
    if re.fullmatch(r"\d+(\.\d+)?m", s):
        return float(s[:-1]) / 12.0
    if re.fullmatch(r"\d+(\.\d+)?", s):
        x = float(s)
        return x if 0 < x <= 100 else None
    return None

def _open_first_xlsx_from_zip(blob):
    if load_workbook is None:
        raise RuntimeError("openpyxl not available")
    with zipfile.ZipFile(io.BytesIO(blob)) as zf:
        candidates = [n for n in zf.namelist() if n.lower().endswith((".xlsx", ".xlsm"))]
        if not candidates:
            raise RuntimeError("no xlsx/xlsm member inside zip")
        return io.BytesIO(zf.read(candidates[0]))

def _find_header_row(ws, min_numeric_headers=6, max_scan_rows=40, max_scan_cols=160):
    for r in range(1, min(ws.max_row, max_scan_rows) + 1):
        nums = []
        for c in range(1, min(ws.max_column, max_scan_cols) + 1):
            yrs = _to_years(ws.cell(r, c).value)
            nums.append(yrs)
        count = sum(1 for x in nums if x is not None)
        if count >= min_numeric_headers:
            return r
    return None

def _find_date_col(ws, header_row, max_scan_cols=12, probe_rows=20):
    best = (None, -1)
    for c in range(1, min(ws.max_column, max_scan_cols) + 1):
        score = 0
        for r in range(header_row + 1, min(ws.max_row, header_row + probe_rows) + 1):
            if _to_iso_date(ws.cell(r, c).value):
                score += 1
        if score > best[1]:
            best = (c, score)
    return best[0]

def _extract_curve_rows_from_sheet(ws, want_map):
    """
    Generic parser for wide curve sheets:
      date | 0.5 | 1 | 2 | 3 | 5 | ...
    """
    header_row = _find_header_row(ws)
    if not header_row:
        return []

    date_col = _find_date_col(ws, header_row)
    if not date_col:
        return []

    numeric_headers = {}
    for c in range(1, ws.max_column + 1):
        yrs = _to_years(ws.cell(header_row, c).value)
        if yrs is not None:
            numeric_headers[c] = yrs

    if len(numeric_headers) < 4:
        return []

    target_cols = {}
    for label, yrs in want_map.items():
        ranked = sorted(numeric_headers.items(), key=lambda kv: abs(kv[1] - yrs))
        if not ranked:
            continue
        c, found_yrs = ranked[0]
        if abs(found_yrs - yrs) <= (0.08 if yrs <= 2 else 0.6):
            target_cols[label] = c

    if len(target_cols) < 3:
        return []

    rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        d = _to_iso_date(ws.cell(r, date_col).value)
        if not d:
            continue
        yd = {}
        for label, c in target_cols.items():
            yd[label] = validate_yield(_as_float(ws.cell(r, c).value))
        if any(v is not None for v in yd.values()):
            rows.append({"date": d, "yields": yd})

    rows.sort(key=lambda x: x["date"], reverse=True)
    return rows

def _load_workbook_from_bytes(blob):
    if load_workbook is None:
        raise RuntimeError("openpyxl not available")
    return load_workbook(io.BytesIO(blob), data_only=True, read_only=True)

def _strip_html(s):
    return " ".join(re.sub(r"<[^>]+>", " ", s).split())

def scrape_investing_yield(url_path):
    try:
        html = get(f"https://www.investing.com{url_path}", timeout=7)
        for pat in [
            r'data-test="instrument-price-last"[^>]*>([\d.]+)<',
            r'class="text-5xl[^"]*"[^>]*>([\d.]+)<',
            r'class="text-2xl[^"]*"[^>]*>([\d.]+)<',
            r'"last":\s*([\d.]+)',
            r'"last_numeric":\s*([\d.]+)',
        ]:
            m = re.search(pat, html)
            if m:
                v = float(m.group(1))
                if 0 < v < 20:
                    return v
    except Exception:
        pass
    return None

def _scrape_tenors(tenor_path_map, max_workers=4):
    """Scrape multiple Investing.com yield paths in parallel. Returns {tenor: value_or_None}."""
    results = {}
    with ThreadPoolExecutor(max_workers=max_workers) as ex:
        fut_map = {ex.submit(scrape_investing_yield, path): tenor for tenor, path in tenor_path_map.items()}
        for fut in as_completed(fut_map):
            tenor = fut_map[fut]
            try:
                results[tenor] = validate_yield(fut.result())
            except Exception:
                results[tenor] = None
    return results

def scrape_commodity_spot(url_path):
    try:
        html = get(f"https://www.investing.com{url_path}", timeout=7)
        for pat in [
            r'data-test="instrument-price-last"[^>]*>([\d,]+\.?\d*)<',
            r'class="text-5xl[^"]*"[^>]*>([\d,]+\.?\d*)<',
            r'"last":\s*([\d.]+)',
            r'"last_numeric":\s*([\d.]+)',
        ]:
            m = re.search(pat, html)
            if m:
                v = float(m.group(1).replace(",", ""))
                if v > 0:
                    return round(v, 2)
    except Exception:
        pass
    return None

def scrape_fx_spot(url_path):
    try:
        html = get(f"https://www.investing.com{url_path}", timeout=8)
        for pat in [
            r'data-test="instrument-price-last"[^>]*>([\d,]+\.?\d*)<',
            r'class="text-5xl[^"]*"[^>]*>([\d,]+\.?\d*)<',
            r'class="text-2xl[^"]*"[^>]*>([\d,]+\.?\d*)<',
            r'"last":\s*([\d.]+)',
            r'"last_numeric":\s*([\d.]+)',
        ]:
            m = re.search(pat, html)
            if m:
                v = float(m.group(1).replace(",", ""))
                if v > 0:
                    return round(v, 4)
    except Exception:
        pass
    return None

def scrape_usdinr_forwards(spot_hint=None):
    """
    Best-effort parse of USD/INR forward rates from Investing.
    Returns tenor map like {"3M": 83.12, "6M": 83.45, "12M": 84.01, "24M": 84.92}.
    """
    out = {}
    try:
        text = _strip_html(get("https://www.investing.com/currencies/usd-inr-forward-rates", timeout=10))
        # Capture tenor tokens followed by nearby decimal quote values.
        # We keep this permissive because page markup can change.
        patterns = {
            "3M": [r"\b3M\b[^0-9]{0,40}([0-9]{2,3}\.[0-9]{2,4})", r"\b3 Month\b[^0-9]{0,40}([0-9]{2,3}\.[0-9]{2,4})"],
            "6M": [r"\b6M\b[^0-9]{0,40}([0-9]{2,3}\.[0-9]{2,4})", r"\b6 Month\b[^0-9]{0,40}([0-9]{2,3}\.[0-9]{2,4})"],
            "12M": [r"\b1Y\b[^0-9]{0,40}([0-9]{2,3}\.[0-9]{2,4})", r"\b12M\b[^0-9]{0,40}([0-9]{2,3}\.[0-9]{2,4})"],
            "24M": [r"\b2Y\b[^0-9]{0,40}([0-9]{2,3}\.[0-9]{2,4})", r"\b24M\b[^0-9]{0,40}([0-9]{2,3}\.[0-9]{2,4})"],
        }
        for tenor, pats in patterns.items():
            for pat in pats:
                m = re.search(pat, text, flags=re.I)
                if m:
                    v = float(m.group(1))
                    # USD/INR outright forwards should be in the same general neighborhood as spot.
                    # Keep wide hard bounds and, when a spot hint is available, reject obviously
                    # misparsed numbers (e.g., forward points / unrelated fields).
                    if not (50 <= v <= 120):
                        continue
                    if spot_hint is not None and spot_hint > 0:
                        max_allowed_diff = 15.0 if tenor == "24M" else 10.0
                        if abs(v - spot_hint) > max_allowed_diff:
                            continue
                    out[tenor] = round(v, 4)
                    break
    except Exception:
        pass
    return out

def scrape_te_last_value(url):
    try:
        text = _strip_html(get(url, timeout=12))
        patterns = [
            r"(?:rose|fell|eased|surged|climbed|hovered|was|traded)\s+(?:to\s+)?([0-9][0-9,]*(?:\.\d+)?)\s+USD",
            r"Actual\s+Chg\s+%Chg\s+[A-Za-z ]+\s+([0-9][0-9,]*(?:\.\d+)?)",
        ]
        for pat in patterns:
            m = re.search(pat, text, flags=re.I)
            if m:
                return round(float(m.group(1).replace(",", "")), 2)
    except Exception:
        pass
    return None

def te_bonds_table(url, code_map):
    """
    Parses the simple bond table visible on TE country bond pages.
    Returns {tenor: {"current":..., "m1":..., "y1":..., "date":...}}.
    """
    try:
        text = _strip_html(get(url, timeout=15))
    except Exception:
        return {}

    out = {}
    for label, te_label in code_map.items():
        pat = re.compile(
            rf"{re.escape(te_label)}\s+([0-9]+(?:\.[0-9]+)?)\s+([+-]?[0-9]+(?:\.[0-9]+)?)%\s+([+-]?[0-9]+(?:\.[0-9]+)?)%\s+([+-]?[0-9]+(?:\.[0-9]+)?)%\s+([A-Za-z]{{3}}/\d{{2}})",
            flags=re.I
        )
        m = pat.search(text)
        if not m:
            continue
        cur = float(m.group(1))
        month_delta = float(m.group(3))
        year_delta = float(m.group(4))
        out[label] = {
            "current": round(cur, 4),
            "m1": round(cur - month_delta, 4),
            "y1": round(cur - year_delta, 4),
            "date": m.group(5),
        }
    return out

def yahoo_price(symbol):
    """Fetch latest close price for a Yahoo Finance symbol (futures, indices, etc.)."""
    try:
        url = f"https://query1.finance.yahoo.com/v8/finance/chart/{symbol}?interval=1d&range=5d"
        data = json.loads(get(url, timeout=6))
        closes = data["chart"]["result"][0]["indicators"]["quote"][0]["close"]
        for v in reversed(closes):
            if v is not None:
                return round(v, 2)
    except Exception:
        pass
    return None

def yahoo_price_at(symbol, days_ago, max_diff_days=5):
    """Fetch close price for a Yahoo Finance symbol approximately N days ago."""
    try:
        target_dt = datetime.utcnow() - timedelta(days=days_ago)
        p1 = int((target_dt - timedelta(days=10)).timestamp())
        p2 = int((target_dt + timedelta(days=3)).timestamp())
        url = f"https://query1.finance.yahoo.com/v8/finance/chart/{symbol}?interval=1d&period1={p1}&period2={p2}"
        data = json.loads(get(url, timeout=6))
        result = data["chart"]["result"][0]
        timestamps = result.get("timestamp", [])
        closes = result["indicators"]["quote"][0]["close"]
        if not timestamps:
            return None, ""
        target_ts = int(target_dt.timestamp())
        pairs = [(ts, c) for ts, c in zip(timestamps, closes) if c is not None]
        if not pairs:
            return None, ""
        best_ts, best_c = min(pairs, key=lambda x: abs(x[0] - target_ts))
        if abs(best_ts - target_ts) > max_diff_days * 86400:
            return None, ""
        return round(best_c, 2), datetime.utcfromtimestamp(best_ts).strftime("%Y-%m-%d")
    except Exception:
        return None, ""

def yahoo_price_near_date(symbol, target_dt, max_diff_days=5):
    try:
        p1 = int((target_dt - timedelta(days=10)).timestamp())
        p2 = int((target_dt + timedelta(days=3)).timestamp())
        url = f"https://query1.finance.yahoo.com/v8/finance/chart/{symbol}?interval=1d&period1={p1}&period2={p2}"
        data = json.loads(get(url, timeout=6))
        result = data["chart"]["result"][0]
        timestamps = result.get("timestamp", [])
        closes = result["indicators"]["quote"][0]["close"]
        if not timestamps:
            return None, ""
        target_ts = int(target_dt.timestamp())
        pairs = [(ts, c) for ts, c in zip(timestamps, closes) if c is not None]
        if not pairs:
            return None, ""
        best_ts, best_c = min(pairs, key=lambda x: abs(x[0] - target_ts))
        if abs(best_ts - target_ts) > max_diff_days * 86400:
            return None, ""
        return round(best_c, 2), datetime.utcfromtimestamp(best_ts).strftime("%Y-%m-%d")
    except Exception:
        return None, ""

# Commodity futures helpers
_MONTH_ABBR = ["jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec"]
_FUT_CODES = {1: "F", 2: "G", 3: "H", 4: "J", 5: "K", 6: "M", 7: "N", 8: "Q", 9: "U", 10: "V", 11: "X", 12: "Z"}
_GOLD_MONTHS = [2, 4, 6, 8, 10, 12]

def _advance_months(m, y, n):
    m += n
    y += (m - 1) // 12
    m = (m - 1) % 12 + 1
    return m, y

def _next_active(m, y, actives):
    for a in actives:
        if a >= m:
            return a, y
    return actives[0], y + 1

def _gold_symbol(months_ahead, base_dt=None):
    base_dt = base_dt or datetime.utcnow()
    m, y = _advance_months(base_dt.month, base_dt.year, months_ahead)
    m, y = _next_active(m, y, _GOLD_MONTHS)
    return f"GC{_FUT_CODES[m]}{y%100:02d}.CMX", f"{_MONTH_ABBR[m-1].capitalize()} {y}"

def _wti_symbol(months_ahead, base_dt=None):
    base_dt = base_dt or datetime.utcnow()
    m, y = _advance_months(base_dt.month, base_dt.year, months_ahead)
    return f"CL{_FUT_CODES[m]}{y%100:02d}.NYM", f"{_MONTH_ABBR[m-1].capitalize()} {y}"

def _brent_symbol(months_ahead, base_dt=None):
    base_dt = base_dt or datetime.utcnow()
    m, y = _advance_months(base_dt.month, base_dt.year, months_ahead)
    return f"BZ{_FUT_CODES[m]}{y%100:02d}.NYM", f"{_MONTH_ABBR[m-1].capitalize()} {y}"

def _usdinr_symbol(months_ahead, base_dt=None):
    """
    INR CME futures-style symbol.
    Falls back to continuous INR=F in fetch logic if contract ticker is unavailable.
    """
    base_dt = base_dt or datetime.utcnow()
    m, y = _advance_months(base_dt.month, base_dt.year, months_ahead)
    return f"INR{_FUT_CODES[m]}{y%100:02d}.CME", f"{_MONTH_ABBR[m-1].capitalize()} {y}"

def _boe_nominal_rows():
    """
    Official Bank of England daily nominal government liability curve archive.
    """
    if load_workbook is None:
        return []
    want = {"1Y": 1.0, "2Y": 2.0, "3Y": 3.0, "5Y": 5.0, "7Y": 7.0, "10Y": 10.0, "15Y": 15.0, "20Y": 20.0, "30Y": 30.0}
    blob = get_bytes("https://www.bankofengland.co.uk/-/media/boe/files/statistics/yield-curves/glcnominalddata.zip", timeout=40, retries=2)
    wb = load_workbook(_open_first_xlsx_from_zip(blob), data_only=True, read_only=True)
    ws = wb["4. spot curve"] if "4. spot curve" in wb.sheetnames else wb[wb.sheetnames[0]]
    return _extract_curve_rows_from_sheet(ws, want)

def _discover_fbil_xlsx_urls():
    urls = []
    try:
        html = get("https://www.fbil.org.in/", timeout=20)
        for href in re.findall(r'href="([^"]+\.(?:xlsx|xlsm|xls))"', html, flags=re.I):
            full = href if href.startswith("http") else f"https://www.fbil.org.in{href}"
            low = full.lower()
            if ("valuation" in low or "gsec" in low or "gs_ec" in low or "yield" in low) and "/uploads/" in low:
                urls.append(full)
    except Exception as e:
        log.warning(f"  FBIL discovery failed: {e}")

    seen = set()
    out = []
    for u in urls:
        if u not in seen:
            seen.add(u)
            out.append(u)
    return out

def _fbil_rows():
    """
    Best-effort workbook parse for latest India G-Sec curve.
    """
    if load_workbook is None:
        return []
    want = {"1Y": 1.0, "2Y": 2.0, "3Y": 3.0, "5Y": 5.0, "7Y": 7.0, "10Y": 10.0, "15Y": 15.0, "20Y": 20.0, "30Y": 30.0}
    for url in _discover_fbil_xlsx_urls():
        try:
            low = url.lower()
            if low.endswith(".xlsx") or low.endswith(".xlsm"):
                wb = _load_workbook_from_bytes(get_bytes(url, timeout=30, retries=1))
            else:
                continue
            for s in wb.sheetnames:
                rows = _extract_curve_rows_from_sheet(wb[s], want)
                if rows:
                    return rows
        except Exception as e:
            log.warning(f"  FBIL parse failed for {url}: {e}")
    return []

# ── 1. UST ──
def fetch_ust():
    log.info("UST: fetching")
    import xml.etree.ElementTree as ET
    ns = {"a": "http://www.w3.org/2005/Atom", "m": "http://schemas.microsoft.com/ado/2007/08/dataservices/metadata", "d": "http://schemas.microsoft.com/ado/2007/08/dataservices"}
    tmap = {"BC_1MONTH": "1M", "BC_3MONTH": "3M", "BC_6MONTH": "6M", "BC_1YEAR": "1Y", "BC_2YEAR": "2Y", "BC_3YEAR": "3Y", "BC_5YEAR": "5Y", "BC_7YEAR": "7Y", "BC_10YEAR": "10Y", "BC_20YEAR": "20Y", "BC_30YEAR": "30Y"}
    tenors = list(tmap.values())

    def parse_year(year):
        raw = get(f"https://home.treasury.gov/resource-center/data-chart-center/interest-rates/pages/xml?data=daily_treasury_yield_curve&field_tdr_date_value={year}")
        rows = []
        for entry in ET.fromstring(raw).findall("a:entry", ns):
            props = entry.find("a:content/m:properties", ns)
            if props is None:
                continue
            de = props.find("d:NEW_DATE", ns)
            if de is None or not de.text:
                continue
            yd = {}
            for xf, tn in tmap.items():
                el = props.find(f"d:{xf}", ns)
                try:
                    yd[tn] = round(float(el.text), 4)
                except Exception:
                    yd[tn] = None
            rows.append({"date": de.text[:10], "yields": yd})
        rows.sort(key=lambda x: x["date"], reverse=True)
        return rows

    now = datetime.utcnow()
    rows = parse_year(now.year)
    assert len(rows) >= 2
    ya_rows = parse_year(now.year - 1)
    # Interpolate 15Y (linear between 10Y and 20Y) for all fetched rows
    for r in rows + ya_rows:
        y = r["yields"]
        y10, y20 = y.get("10Y"), y.get("20Y")
        y["15Y"] = round(y10 + (y20 - y10) * 0.5, 4) if y10 is not None and y20 is not None else None
    tenors = ['1M', '3M', '6M', '1Y', '2Y', '3Y', '5Y', '7Y', '10Y', '15Y', '20Y', '30Y']
    target_ya = (now - timedelta(days=365)).strftime("%Y-%m-%d")
    ya_yields, ya_date = [None] * len(tenors), ""
    if ya_rows:
        best = min(ya_rows, key=lambda r: abs((datetime.strptime(r["date"], "%Y-%m-%d") - datetime.strptime(target_ya, "%Y-%m-%d")).days))
        ya_yields = [best["yields"].get(t) for t in tenors]
        ya_date = best["date"]
    all_rows = rows + ya_rows
    p1m_yields, p1m_date = find_prior_date_yields(all_rows, 30, tenors)
    p3m_yields, p3m_date = find_prior_date_yields(all_rows, 91, tenors)
    log.info(f"  UST 1M ago: {p1m_date}, 3M ago: {p3m_date}")
    write("ust.json", {
        "date": rows[0]["date"],
        "prior_date": rows[1]["date"],
        "source": "US Treasury Daily Par Yield Curve",
        "url": "https://home.treasury.gov/resource-center/data-chart-center/interest-rates",
        "tenors": tenors,
        "yields": [rows[0]["yields"].get(t) for t in tenors],
        "prior_yields": [rows[1]["yields"].get(t) for t in tenors],
        "prior_1m_yields": p1m_yields,
        "prior_1m_date": p1m_date,
        "prior_3m_yields": p3m_yields,
        "prior_3m_date": p3m_date,
        "year_ago_yields": ya_yields,
        "year_ago_date": ya_date
    })
    log.info(f"  UST OK: {rows[0]['date']}")

# ── 2. JGB ──
def fetch_jgb():
    log.info("JGB: fetching")
    want = ["1Y", "2Y", "3Y", "5Y", "7Y", "10Y", "15Y", "20Y", "25Y", "30Y", "40Y"]
    raw = get("https://www.mof.go.jp/english/policy/jgbs/reference/interest_rate/jgbcme.csv")
    lines = raw.split("\n")
    hdr_idx, headers = -1, []
    for i, line in enumerate(lines[:5]):
        if "date" in line.lower():
            hdr_idx = i
            headers = [h.strip().strip('"') for h in line.split(",")]
            break
    assert hdr_idx >= 0
    col = {h.replace(" ", ""): j for j, h in enumerate(headers) if h.replace(" ", "") in want}
    rows = []
    for line in lines[hdr_idx+1:]:
        parts = [p.strip().strip('"') for p in line.split(",")]
        if len(parts) < 10:
            continue
        m = re.match(r"(\d{4})[/-](\d{1,2})[/-](\d{1,2})", parts[0])
        if not m:
            continue
        date_val = f"{m[1]}-{m[2].zfill(2)}-{m[3].zfill(2)}"
        yd = {}
        for t in want:
            if t in col:
                try:
                    yd[t] = round(float(parts[col[t]]), 4)
                except Exception:
                    yd[t] = None
        rows.append({"date": date_val, "yields": yd})
    rows.sort(key=lambda x: x["date"], reverse=True)
    assert len(rows) >= 2
    target_ya = (datetime.utcnow() - timedelta(days=365)).strftime("%Y-%m-%d")
    ya_candidates = [r for r in rows if r["date"] <= target_ya]
    ya_yields, ya_date = [None] * len(want), ""
    if ya_candidates:
        ya_yields = [ya_candidates[0]["yields"].get(t) for t in want]
        ya_date = ya_candidates[0]["date"]
    if not any(v is not None for v in ya_yields):
        fred_ya, fdate = fred_year_ago_10y("IRLTLT01JPM156N")
        if fred_ya:
            ya_yields[want.index("10Y")] = fred_ya
            ya_date = fdate
    p1m_yields, p1m_date = find_prior_date_yields(rows, 30, want)
    p3m_yields, p3m_date = find_prior_date_yields(rows, 91, want)
    log.info(f"  JGB 1M ago: {p1m_date}, 3M ago: {p3m_date}")
    write("jgb.json", {
        "date": rows[0]["date"],
        "prior_date": rows[1]["date"],
        "source": "Ministry of Finance Japan",
        "url": "https://www.mof.go.jp/english/policy/jgbs/reference/interest_rate/",
        "tenors": want,
        "yields": [rows[0]["yields"].get(t) for t in want],
        "prior_yields": [rows[1]["yields"].get(t) for t in want],
        "prior_1m_yields": p1m_yields,
        "prior_1m_date": p1m_date,
        "prior_3m_yields": p3m_yields,
        "prior_3m_date": p3m_date,
        "year_ago_yields": ya_yields,
        "year_ago_date": ya_date
    })
    log.info(f"  JGB OK: {rows[0]['date']}")

# ── 3. GILT ──
GILT_INV = {"1Y": "/rates-bonds/uk-1-year-bond-yield", "2Y": "/rates-bonds/uk-2-year-bond-yield", "3Y": "/rates-bonds/uk-3-year-bond-yield", "5Y": "/rates-bonds/uk-5-year-bond-yield", "7Y": "/rates-bonds/uk-7-year-bond-yield", "10Y": "/rates-bonds/uk-10-year-bond-yield", "15Y": "/rates-bonds/uk-15-year-bond-yield", "20Y": "/rates-bonds/uk-20-year-bond-yield", "30Y": "/rates-bonds/uk-30-year-bond-yield"}

def fetch_gilt():
    log.info("GILT: fetching")
    tenors = ["1Y", "2Y", "3Y", "5Y", "7Y", "10Y", "15Y", "20Y", "30Y"]

    rows = []
    try:
        rows = _boe_nominal_rows()
        if rows:
            log.info(f"  GILT official BoE rows: {len(rows)}")
    except Exception as e:
        log.warning(f"  GILT BoE failed: {e}")

    current = {}
    if rows:
        current = rows[0]["yields"].copy()
        date_str = rows[0]["date"]
        prior_yields = [rows[1]["yields"].get(t) for t in tenors] if len(rows) > 1 else [None] * len(tenors)
        prior_date = rows[1]["date"] if len(rows) > 1 else ""
        p1m_yields, p1m_date = find_prior_date_yields(rows, 30, tenors)
        p3m_yields, p3m_date = find_prior_date_yields(rows, 91, tenors)
        ya_yields, ya_date = find_prior_date_yields(rows, 365, tenors, max_diff_days=21)
        source = "Bank of England daily government liability curve (nominal)"
        source_url = "https://www.bankofengland.co.uk/-/media/boe/files/statistics/yield-curves/glcnominalddata.zip"
    else:
        scraped = _scrape_tenors(GILT_INV)
        current.update({k: v for k, v in scraped.items() if v is not None})
        te = te_bonds_table(
            "https://tradingeconomics.com/united-kingdom/government-bond-yield",
            {
                "1Y": "UK 52W",
                "2Y": "UK 2Y",
                "3Y": "UK 3Y",
                "5Y": "UK 5Y",
                "7Y": "UK 7Y",
                "10Y": "UK 10Y",
                "15Y": "UK 15Y",
                "20Y": "UK 20Y",
                "30Y": "UK 30Y",
            }
        )
        for t in tenors:
            if current.get(t) is None and te.get(t):
                current[t] = validate_yield(te[t]["current"])

        try:
            obs = fred_csv("IRLTLT01GBM156N", start="2024-01-01")
            if obs and current.get("10Y") is None:
                current["10Y"] = validate_yield(obs[0]["value"])
        except Exception:
            pass

        current = interpolate_curve(current, tenors)

        last = load_last_gilt()
        if last:
            last_y = dict(zip(last.get("tenors", []), last.get("yields", [])))
            for t in tenors:
                if current.get(t) is None and last_y.get(t) is not None:
                    current[t] = last_y[t]

        p1m_yields = [te.get(t, {}).get("m1") for t in tenors]
        p1m_date = "TE month delta reconstruction"
        p3m_yields, p3m_date = history_lookup("gilt", 91, tenors)
        ya_yields = [te.get(t, {}).get("y1") for t in tenors]
        ya_date = "TE year delta reconstruction"
        prior_yields = [None] * len(tenors)
        prior_date = ""
        date_str = datetime.utcnow().strftime("%Y-%m-%d")
        source = "Investing.com / TradingEconomics / FRED / cache"
        source_url = "https://www.investing.com/rates-bonds/uk-government-bonds"

    valid_count = sum(1 for v in current.values() if v is not None)
    if valid_count < 3:
        raise Exception(f"GILT: insufficient data ({valid_count} tenors)")

    append_curve_history("gilt", date_str, tenors, [current.get(t) for t in tenors], source)

    write("gilt.json", {
        "date": date_str,
        "prior_date": prior_date,
        "source": source,
        "url": source_url,
        "tenors": tenors,
        "yields": [current.get(t) for t in tenors],
        "prior_yields": prior_yields,
        "prior_1m_yields": p1m_yields,
        "prior_1m_date": p1m_date,
        "prior_3m_yields": p3m_yields,
        "prior_3m_date": p3m_date,
        "year_ago_yields": ya_yields,
        "year_ago_date": ya_date,
        "note": "Official BoE daily archive first; then market scrapes; then local history / cache."
    })
    log.info(f"  GILT OK: {valid_count} tenors")

# ── 4. EUR ──
def fetch_eur():
    log.info("EUR: fetching")
    ecb_map = {"1Y": "SR_1Y", "2Y": "SR_2Y", "3Y": "SR_3Y", "5Y": "SR_5Y", "7Y": "SR_7Y", "10Y": "SR_10Y", "15Y": "SR_15Y", "20Y": "SR_20Y", "30Y": "SR_30Y"}
    tenors = list(ecb_map.keys())
    results = {}
    for tn, sk in ecb_map.items():
        try:
            raw = get(f"https://data-api.ecb.europa.eu/service/data/YC/B.U2.EUR.4F.G_N_A.SV_C_YM.{sk}?lastNObservations=70&format=csvdata", timeout=15)
            lines = raw.strip().split("\n")
            if len(lines) < 2:
                continue
            header = lines[0].split(",")
            oi = next((i for i, h in enumerate(header) if "OBS_VALUE" in h), -1)
            ti = next((i for i, h in enumerate(header) if "TIME_PERIOD" in h), -1)
            if oi < 0:
                continue
            obs = []
            for line in lines[1:]:
                p = line.split(",")
                try:
                    obs.append({"date": p[ti].strip('"'), "value": round(float(p[oi]), 4)})
                except Exception:
                    pass
            obs.sort(key=lambda x: x["date"], reverse=True)
            if obs:
                results[tn] = {"value": obs[0]["value"], "prior": obs[1]["value"] if len(obs) > 1 else None, "date": obs[0]["date"], "all_obs": obs}
        except Exception:
            pass
    assert results
    latest = max(r["date"] for r in results.values())

    def ecb_find_prior(days_ago):
        target = (datetime.utcnow() - timedelta(days=days_ago)).strftime("%Y-%m-%d")
        out = []
        d_used = ""
        for tn in tenors:
            obs_list = results.get(tn, {}).get("all_obs", [])
            if not obs_list:
                out.append(None)
                continue
            best = min(obs_list, key=lambda o: abs((datetime.strptime(o["date"], "%Y-%m-%d") - datetime.strptime(target, "%Y-%m-%d")).days))
            diff = abs((datetime.strptime(best["date"], "%Y-%m-%d") - datetime.strptime(target, "%Y-%m-%d")).days)
            if diff > 14:
                out.append(None)
            else:
                out.append(best["value"])
                d_used = best["date"]
        return out, d_used

    p1m_yields, p1m_date = ecb_find_prior(30)
    p3m_yields, p3m_date = ecb_find_prior(91)
    log.info(f"  EUR 1M ago: {p1m_date}, 3M ago: {p3m_date}")
    write("eur.json", {
        "date": latest,
        "prior_date": "",
        "source": "ECB SDW (EUR AAA Govt — EIOPA proxy)",
        "url": "https://data.ecb.europa.eu/",
        "tenors": tenors,
        "yields": [results.get(t, {}).get("value") for t in tenors],
        "prior_yields": [results.get(t, {}).get("prior") for t in tenors],
        "prior_1m_yields": p1m_yields,
        "prior_1m_date": p1m_date,
        "prior_3m_yields": p3m_yields,
        "prior_3m_date": p3m_date,
        "year_ago_yields": [None] * len(tenors),
        "year_ago_date": "",
        "note": "EUR AAA govt curve proxy. Actual EIOPA RFR includes UFR extrapolation."
    })
    log.info(f"  EUR OK: {latest}")

# ── 5. INDIA ──
INDIA_INV = {"1Y": "/rates-bonds/india-1-year-bond-yield", "2Y": "/rates-bonds/india-2-year-bond-yield", "3Y": "/rates-bonds/india-3-year-bond-yield", "5Y": "/rates-bonds/india-5-year-bond-yield", "7Y": "/rates-bonds/india-7-year-bond-yield", "10Y": "/rates-bonds/india-10-year-bond-yield", "15Y": "/rates-bonds/india-15-year-bond-yield", "20Y": "/rates-bonds/india-20-year-bond-yield", "30Y": "/rates-bonds/india-30-year-bond-yield"}

def fetch_india():
    log.info("INDIA: fetching")
    tenors = ["1Y", "2Y", "3Y", "5Y", "7Y", "10Y", "15Y", "20Y", "30Y"]

    current = {}
    date_str = datetime.utcnow().strftime("%Y-%m-%d")
    source = ""
    source_url = ""

    rows = []
    try:
        rows = _fbil_rows()
        if rows:
            current = rows[0]["yields"].copy()
            date_str = rows[0]["date"]
            source = "FBIL GOI prices / par yield workbook"
            source_url = "https://www.fbil.org.in/"
            log.info(f"  INDIA FBIL rows: {len(rows)}")
    except Exception as e:
        log.warning(f"  INDIA FBIL failed: {e}")

    if not current:
        scraped = _scrape_tenors(INDIA_INV)
        current.update({k: v for k, v in scraped.items() if v is not None})
        source = "Investing.com / TradingEconomics / FRED / cache"
        source_url = "https://www.investing.com/rates-bonds/india-government-bonds"

    te = te_bonds_table(
        "https://tradingeconomics.com/india/government-bond-yield",
        {
            "1Y": "India 52W",
            "2Y": "India 2Y",
            "3Y": "India 3Y",
            "5Y": "India 5Y",
            "7Y": "India 7Y",
            "10Y": "India 10Y",
            "15Y": "India 15Y",
            "20Y": "India 20Y",
            "30Y": "India 30Y",
        }
    )

    for t in tenors:
        if current.get(t) is None and te.get(t):
            current[t] = validate_yield(te[t]["current"])

    try:
        obs = fred_csv("INDIRLTLT01STM", start="2024-01-01")
        if obs and current.get("10Y") is None:
            current["10Y"] = validate_yield(obs[0]["value"])
    except Exception:
        pass

    current = interpolate_curve(current, tenors)

    last = load_last_india()
    if last:
        last_y = dict(zip(last.get("tenors", []), last.get("yields", [])))
        for t in tenors:
            if current.get(t) is None and last_y.get(t) is not None:
                current[t] = last_y[t]

    valid_count = sum(1 for v in current.values() if v is not None)
    if valid_count < 3:
        raise Exception(f"INDIA: insufficient data ({valid_count} tenors)")

    if rows and len(rows) > 2:
        p1m_yields, p1m_date = find_prior_date_yields(rows, 30, tenors)
        p3m_yields, p3m_date = find_prior_date_yields(rows, 91, tenors)
        ya_yields, ya_date = find_prior_date_yields(rows, 365, tenors, max_diff_days=21)
    else:
        p1m_yields = [te.get(t, {}).get("m1") for t in tenors]
        p1m_date = "TE month delta reconstruction"
        p3m_yields, p3m_date = history_lookup("india", 91, tenors)
        ya_yields = [te.get(t, {}).get("y1") for t in tenors]
        ya_date = "TE year delta reconstruction"
        # Only call FRED for 10Y priors that TE didn't cover (avoids expensive
        # timeout chains when FRED is under load)
        idx10 = tenors.index("10Y")
        need_p1m = p1m_yields[idx10] is None
        need_p3m = p3m_yields[idx10] is None
        need_ya  = ya_yields[idx10] is None
        if need_p1m or need_p3m or need_ya:
            try:
                if need_p1m:
                    p1m_val, p1m_d = fred_prior_single("INDIRLTLT01STM", 30)
                    if p1m_val is not None:
                        p1m_yields[idx10] = p1m_val
                        p1m_date = p1m_d or p1m_date
                if need_p3m:
                    p3m_val, p3m_d = fred_prior_single("INDIRLTLT01STM", 91)
                    if p3m_val is not None:
                        p3m_yields[idx10] = p3m_val
                        p3m_date = p3m_d or p3m_date
                if need_ya:
                    ya_val, ya_d = fred_year_ago_10y("INDIRLTLT01STM")
                    if ya_val is not None:
                        ya_yields[idx10] = ya_val
                        ya_date = ya_d or ya_date
            except Exception:
                pass

    append_curve_history("india", date_str, tenors, [current.get(t) for t in tenors], source)

    write("india.json", {
        "date": date_str,
        "source": source,
        "url": source_url,
        "tenors": tenors,
        "yields": [current.get(t) for t in tenors],
        "prior_yields": [None] * len(tenors),
        "prior_1m_yields": p1m_yields,
        "prior_1m_date": p1m_date,
        "prior_3m_yields": p3m_yields,
        "prior_3m_date": p3m_date,
        "year_ago_yields": ya_yields,
        "year_ago_date": ya_date,
        "note": "FBIL latest first, then Investing/TE/FRED, then local history / cache. India 3M curve is exact once local history accumulates."
    })
    log.info(f"  INDIA OK: {valid_count} tenors")

# ── 6. CREDIT ──
def fetch_credit():
    log.info("CREDIT: fetching")
    series = {
        "ig": "BAMLC0A0CM",
        "aaa": "BAMLC0A1CAAA",
        "aa": "BAMLC0A2CAA",
        "a": "BAMLC0A3CA",
        "bbb": "BAMLC0A4CBBB",
        "hy": "BAMLH0A0HYM2",
        "bb": "BAMLH0A1HYBB",
        "b": "BAMLH0A2HYB",
        "ccc": "BAMLH0A3HYC",
    }
    names = {
        "ig": "US IG",
        "aaa": "US AAA",
        "aa": "US AA",
        "a": "US A",
        "bbb": "US BBB",
        "hy": "US HY",
        "bb": "US BB",
        "b": "US B",
        "ccc": "US CCC+",
    }
    buckets = {
        "ig": "IG",
        "aaa": "AAA",
        "aa": "AA",
        "a": "A",
        "bbb": "BBB",
        "hy": "HY",
        "bb": "BB",
        "b": "B",
        "ccc": "CCC",
    }

    def build_row(key, obs, source_tag):
        curr = round(obs[0]["value"] * 100)
        prev = round(obs[1]["value"] * 100) if len(obs) > 1 else curr
        return {
            "name": names[key],
            "spread": curr,
            "prior": prev,
            "bucket": buckets[key],
            "date": obs[0]["date"],
            "source": source_tag,
        }

    spreads = {}
    latest_date = ""

    # ── Stage 1: single bulk FRED call (no retries) ──
    multi = fred_multi_csv(list(series.values()), start="2024-01-01", retries=0)
    for key, sid in series.items():
        obs = multi.get(sid, [])
        if obs:
            spreads[key] = build_row(key, obs, "fred_multi")
            if obs[0]["date"] > latest_date:
                latest_date = obs[0]["date"]
            log.info(f"  Credit {key}: {spreads[key]['spread']}bp (multi)")
        else:
            log.warning(f"  Credit {key}: missing from bulk fetch")

    # ── Stage 2: alternative FRED download endpoint for missing series (parallel) ──
    missing = [key for key in series if key not in spreads]
    if missing:
        log.info(f"  Credit: trying download endpoint for: {missing}")
        with ThreadPoolExecutor(max_workers=min(4, len(missing))) as ex:
            fut_map = {ex.submit(fred_download_csv, series[key], "2024-01-01"): key for key in missing}
            for fut in as_completed(fut_map):
                key = fut_map[fut]
                try:
                    obs = fut.result()
                except Exception:
                    obs = []
                if obs:
                    spreads[key] = build_row(key, obs, "fred_download")
                    if obs[0]["date"] > latest_date:
                        latest_date = obs[0]["date"]
                    log.info(f"  Credit {key}: {spreads[key]['spread']}bp (download)")
                else:
                    log.warning(f"  Credit {key}: download endpoint also failed")

    # ── Stage 3: cache fallback for anything still missing ──
    last = load_last_credit()
    if last:
        last_spreads = last.get("spreads", {})
        for key in series:
            if key not in spreads and key in last_spreads:
                cached = dict(last_spreads[key])
                cached["source"] = "cache"
                spreads[key] = cached
                latest_date = max(latest_date, cached.get("date", ""))
                log.warning(f"  Credit {key}: using cache fallback")

    if not spreads:
        raise Exception("CREDIT: no series")

    any_cached = any(spreads[k].get("source") == "cache" for k in spreads)
    note = (
        "Single bulk FRED fetch; missing series retried via FRED download endpoint; "
        "cache fallback used for unavailable series. Values are option-adjusted spreads in basis points."
        + (" Some series are from cache." if any_cached else "")
    )

    write("credit.json", {
        "date": latest_date,
        "source": "FRED / ICE BofA Indices",
        "url": "https://fred.stlouisfed.org/release?rid=209",
        "spreads": spreads,
        "note": note,
    })
    log.info(f"  CREDIT OK: {latest_date}, {len(spreads)} series")

# ── 7. SOFR ──
def fetch_sofr():
    log.info("SOFR: fetching from NY Fed API")
    rates = {}
    history = []
    latest_date = ""

    try:
        url = "https://markets.newyorkfed.org/api/rates/secured/sofr/last/270.json"
        raw = get(url, timeout=15)
        data = json.loads(raw)
        sofr_data = data.get("refRates", [])
        if sofr_data:
            sofr_data.sort(key=lambda x: x.get("effectiveDate", ""), reverse=True)
            latest = sofr_data[0]
            prior = sofr_data[1] if len(sofr_data) > 1 else sofr_data[0]
            latest_date = latest.get("effectiveDate", "")
            rates["SOFR"] = {
                "name": "SOFR (Daily)",
                "desc": "Secured Overnight Financing Rate",
                "rate": round(float(latest.get("percentRate", 0)), 4),
                "prior": round(float(prior.get("percentRate", 0)), 4),
                "date": latest_date,
                "volume": latest.get("volumeInBillions"),
                "percentile_25": latest.get("percentPercentile25"),
                "percentile_75": latest.get("percentPercentile75"),
            }
            log.info(f"  SOFR daily: {rates['SOFR']['rate']}% ({latest_date})")
            for d in sofr_data:
                try:
                    history.append({"date": d["effectiveDate"], "rate": round(float(d["percentRate"]), 4)})
                except Exception:
                    pass
            history.sort(key=lambda x: x["date"])
    except Exception as e:
        log.warning(f"  SOFR NY Fed API: {e}")

    try:
        url = "https://markets.newyorkfed.org/api/rates/secured/sofr/last/1.json?productType=sofrAverage"
        raw = get(url, timeout=15)
        data = json.loads(raw)
        for item in data.get("refRates", []):
            avg_type = item.get("averagingMethod", "")
            if "30" in avg_type:
                rates["30D_AVG"] = {"name": "SOFR 30-Day Avg", "desc": "30-day compounded average", "rate": round(float(item.get("percentRate", 0)), 4), "prior": None, "date": item.get("effectiveDate", "")}
                log.info(f"  SOFR 30D: {rates['30D_AVG']['rate']}%")
            elif "90" in avg_type:
                rates["90D_AVG"] = {"name": "SOFR 90-Day Avg", "desc": "90-day compounded average", "rate": round(float(item.get("percentRate", 0)), 4), "prior": None, "date": item.get("effectiveDate", "")}
                log.info(f"  SOFR 90D: {rates['90D_AVG']['rate']}%")
            elif "180" in avg_type:
                rates["180D_AVG"] = {"name": "SOFR 180-Day Avg", "desc": "180-day compounded average", "rate": round(float(item.get("percentRate", 0)), 4), "prior": None, "date": item.get("effectiveDate", "")}
                log.info(f"  SOFR 180D: {rates['180D_AVG']['rate']}%")
    except Exception as e:
        log.warning(f"  SOFR averages: {e}")

    if not rates:
        log.info("  SOFR: trying FRED fallback")
        for key, sid, name, desc in [
            ("SOFR", "SOFR", "SOFR (Daily)", "Secured Overnight Financing Rate"),
            ("30D_AVG", "SOFR30DAYAVG", "SOFR 30-Day Avg", "30-day compounded average"),
            ("90D_AVG", "SOFR90DAYAVG", "SOFR 90-Day Avg", "90-day compounded average"),
            ("180D_AVG", "SOFR180DAYAVG", "SOFR 180-Day Avg", "180-day compounded average"),
        ]:
            obs = fred_csv(sid, start="2025-01-01", retries=1)
            if obs:
                rates[key] = {"name": name, "desc": desc, "rate": round(obs[0]["value"], 4), "prior": round(obs[1]["value"], 4) if len(obs) > 1 else None, "date": obs[0]["date"]}
                if key == "SOFR":
                    for o in obs[:270]:
                        history.append({"date": o["date"], "rate": round(o["value"], 4)})
                    history.sort(key=lambda x: x["date"])
                    if obs[0]["date"] > latest_date:
                        latest_date = obs[0]["date"]
            time.sleep(1)

    assert rates, "SOFR: no data from NY Fed or FRED"

    ya_rate, ya_date = None, ""
    try:
        target = (datetime.utcnow() - timedelta(days=365)).strftime("%Y-%m-%d")
        url = f"https://markets.newyorkfed.org/api/rates/secured/sofr/search.json?startDate={(datetime.utcnow()-timedelta(days=370)).strftime('%Y-%m-%d')}&endDate={(datetime.utcnow()-timedelta(days=360)).strftime('%Y-%m-%d')}"
        raw = get(url, timeout=10)
        data = json.loads(raw)
        ya_data = data.get("refRates", [])
        if ya_data:
            best = min(ya_data, key=lambda x: abs((datetime.strptime(x["effectiveDate"], "%Y-%m-%d") - datetime.strptime(target, "%Y-%m-%d")).days))
            ya_rate = round(float(best["percentRate"]), 4)
            ya_date = best["effectiveDate"]
    except Exception:
        pass
    if ya_rate is None:
        ya_rate, ya_date = fred_year_ago_10y("SOFR")

    log.info(f"  SOFR year-ago: {ya_rate}% ({ya_date})")

    # Merge UST 3M and 1Y into history records
    try:
        one_yr_ago = (datetime.utcnow() - timedelta(days=380)).strftime("%Y-%m-%d")
        ust_raw = fred_multi_csv(["DGS3MO", "DGS1"], start=one_yr_ago)
        ust_3m_map = {o["date"]: round(o["value"], 4) for o in (ust_raw.get("DGS3MO") or [])}
        ust_1y_map = {o["date"]: round(o["value"], 4) for o in (ust_raw.get("DGS1")   or [])}
        for rec in history:
            rec["ust_3m"] = ust_3m_map.get(rec["date"])
            rec["ust_1y"] = ust_1y_map.get(rec["date"])
        log.info(f"  UST 3M/1Y merged: {len(ust_3m_map)} / {len(ust_1y_map)} observations")
    except Exception as e:
        log.warning(f"  UST history merge: {e}")

    # Term SOFR (forward-looking CME reference rates via FRED)
    term_rates = {}
    try:
        tr_start = (datetime.utcnow() - timedelta(days=7)).strftime("%Y-%m-%d")
        tr_raw = fred_multi_csv(["SOFRTERM1M", "SOFRTERM3M", "SOFRTERM6M", "SOFRTERM1Y"], start=tr_start)
        term_map = {"SOFRTERM1M": "1M", "SOFRTERM3M": "3M", "SOFRTERM6M": "6M", "SOFRTERM1Y": "1Y"}
        term_labels = {"1M": "Term SOFR 1M", "3M": "Term SOFR 3M", "6M": "Term SOFR 6M", "1Y": "Term SOFR 1Y"}
        for sid, key in term_map.items():
            obs = tr_raw.get(sid) or []
            if obs:
                cur = obs[0]
                prior = obs[1] if len(obs) > 1 else None
                term_rates[key] = {
                    "name": term_labels[key],
                    "rate": round(cur["value"], 4),
                    "prior": round(prior["value"], 4) if prior else None,
                    "date": cur["date"],
                }
        log.info(f"  Term SOFR: {list(term_rates.keys())}")
    except Exception as e:
        log.warning(f"  Term SOFR: {e}")

    write("sofr.json", {
        "date": latest_date,
        "source": "NY Fed / FRED",
        "url": "https://www.newyorkfed.org/markets/reference-rates/sofr",
        "rates": rates,
        "history": history,
        "year_ago": {"rate": ya_rate, "date": ya_date},
        "term_rates": term_rates,
        "note": "Published daily by NY Fed at ~8:00 AM ET. Averages are backward-looking compounded. Term SOFR via CME/FRED."
    })
    log.info(f"  SOFR OK: {latest_date}")


def _parse_iso_prefix_date_from_url(url):
    m = re.search(r"/documents/(\d{4}-\d{2}-\d{2})-", url)
    if m:
        try:
            return datetime.strptime(m.group(1), "%Y-%m-%d")
        except Exception:
            return None
    return None

def _quarter_end_from_date(dt):
    q = (dt.month - 1) // 3 + 1
    if q == 1:
        return datetime(dt.year, 3, 31)
    if q == 2:
        return datetime(dt.year, 6, 30)
    if q == 3:
        return datetime(dt.year, 9, 30)
    return datetime(dt.year, 12, 31)

def _quarter_label(dt):
    q = (dt.month - 1) // 3 + 1
    return f"{dt.year} Q{q}"

def _bma_filename_title_date(url):
    return extract_date_from_url(url)

def _bma_normalize_tenor(v):
    if v is None:
        return None
    s = str(v).strip().lower()
    m = re.match(r"(\d+)\s*year", s)
    if m:
        return f"{int(m.group(1))}Y"
    m = re.match(r"(\d+)\s*month", s)
    if m:
        return f"{int(m.group(1))}M"
    return None

def _bma_currency_code(name):
    mapping = {
        "US": "USD",
        "UK": "GBP",
        "Switzerland": "CHF",
        "Canada": "CAD",
        "Japan": "JPY",
        "Australia": "AUD",
        "New Zealand": "NZD",
        "Euro Area": "EUR",
        "EUR": "EUR",
        "Europe": "EUR",
    }
    s = str(name).strip()
    return mapping.get(s, s.upper()[:3])

BMA_KNOWN_DISCOUNT_FILES = {
    "2025-12-31": "https://cdn.bma.bm/documents/2026-01-15-11-20-15-Discount-Rates.-31-December-2025.xlsx",
    "2025-09-30": "https://cdn.bma.bm/documents/2025-10-22-15-18-14-Discount-Rates.-30-September-2025..xlsx",
    "2025-06-30": "https://cdn.bma.bm/documents/2025-07-18-10-22-07-Discount-Rates.-30-June-2025..xlsx",
    "2025-03-31": "https://cdn.bma.bm/documents/2025-04-15-14-44-04-Discount-Rates.-31-March-2025..xlsx",
    "2024-12-31": "https://cdn.bma.bm/documents/2025-01-17-16-39-24-Discount-Rates.-31-December-2024..xlsx",
}

BMA_DOC_PAGES = [
    "https://www.bma.bm/documents-centre/documents-reporting-forms-and-guidelines/documents-insurance",
    "https://www.bma.bm/documents-centre/documents-reporting-forms-and-guidelines",
    "https://www.bma.bm/document-centre/reporting-forms-and-guidelines-insurance",
    "https://www.bma.bm/document-centre/reporting-forms-and-guidelines",
]

def _bma_discover_discount_files():
    """
    Discover BMA discount-rate workbooks/attachments from several BMA document-centre pages.
    Falls back to a small recent-quarter map because the site markup and pagination are inconsistent.
    """
    entries = {}

    def upsert(as_of_dt, uploaded_dt, url, source_page):
        if not as_of_dt or not url:
            return
        key = as_of_dt.strftime("%Y-%m-%d")
        cur = entries.get(key)
        candidate = {
            "as_of_dt": as_of_dt,
            "as_of": as_of_dt.strftime("%d %B %Y").lstrip("0"),
            "uploaded_dt": uploaded_dt,
            "uploaded_on": uploaded_dt.strftime("%d %B %Y").lstrip("0") if uploaded_dt else "",
            "url": url,
            "source_page": source_page,
        }
        if (cur is None) or ((uploaded_dt or datetime.min) > (cur.get("uploaded_dt") or datetime.min)):
            entries[key] = candidate

    href_pat = re.compile(r'href="([^"]*(?:Discount[-\s_.]*Rates|discount[-\s_.]*rates)[^"]*\.(?:xlsx|xlsm|xls|pdf))"', re.I)
    title_pat = re.compile(r"Discount\s+Rates\.?\s*(\d{1,2}\s+\w+\s+\d{4})", re.I)
    upload_pat = re.compile(r"Uploaded on\s+(\d{1,2}\s+\w+\s+\d{4})", re.I)

    for page in BMA_DOC_PAGES:
        try:
            html = get(page, timeout=20)
        except Exception as e:
            log.warning(f"  BMA page fetch failed {page}: {e}")
            continue

        # Strong block-level parse: title + uploaded date + href in proximity.
        block_pat = re.compile(
            r"Discount\s+Rates\.?\s*(\d{1,2}\s+\w+\s+\d{4}).{0,1200}?Uploaded on\s+(\d{1,2}\s+\w+\s+\d{4}).{0,1200}?href=\"([^\"]*(?:Discount[-\s_.]*Rates|discount[-\s_.]*rates)[^\"]*\.(?:xlsx|xlsm|xls|pdf))\"",
            re.I | re.S
        )
        for asof_raw, uploaded_raw, href_raw in block_pat.findall(html):
            as_of_dt = parse_d(asof_raw)
            uploaded_dt = parse_d(uploaded_raw)
            url = href_raw if href_raw.startswith("http") else f"https://www.bma.bm{href_raw}"
            upsert(as_of_dt, uploaded_dt, url, page)

        # Generic href scan with context fallback.
        for m in href_pat.finditer(html):
            href_raw = m.group(1)
            url = href_raw if href_raw.startswith("http") else f"https://www.bma.bm{href_raw}"
            context = html[max(0, m.start()-1200):m.end()+1200]
            asof_dt = None
            uploaded_dt = None
            mt = title_pat.search(context)
            mu = upload_pat.search(context)
            if mt:
                asof_dt = parse_d(mt.group(1))
            if mu:
                uploaded_dt = parse_d(mu.group(1))
            if asof_dt is None:
                asof_dt = _bma_filename_title_date(url)
            if uploaded_dt is None:
                uploaded_dt = _parse_iso_prefix_date_from_url(url)
            upsert(asof_dt, uploaded_dt, url, page)

    # Recent-quarter hard fallback for resilience.
    for k, url in BMA_KNOWN_DISCOUNT_FILES.items():
        as_of_dt = datetime.strptime(k, "%Y-%m-%d")
        uploaded_dt = _parse_iso_prefix_date_from_url(url)
        upsert(as_of_dt, uploaded_dt, url, "known_fallback")

    out = list(entries.values())
    out.sort(key=lambda x: (x["as_of_dt"], x["uploaded_dt"] or datetime.min), reverse=True)
    return out

def _bma_extract_table(ws, title_text):
    title_cell = None
    for r in range(1, min(ws.max_row, 25) + 1):
        for c in range(1, min(ws.max_column, 30) + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and title_text.lower() in v.lower():
                title_cell = (r, c)
                break
        if title_cell:
            break
    if not title_cell:
        return {}

    title_row, title_col = title_cell
    header_row = title_row + 1
    maturity_col = title_col
    currencies = []
    c = maturity_col + 1
    while c <= ws.max_column:
        hv = ws.cell(header_row, c).value
        if hv in (None, ""):
            break
        currencies.append((c, _bma_currency_code(hv)))
        c += 1

    table = {ccy: {} for _, ccy in currencies}
    r = header_row + 1
    blank_streak = 0
    while r <= ws.max_row:
        tenor = _bma_normalize_tenor(ws.cell(r, maturity_col).value)
        if not tenor:
            blank_streak += 1
            if blank_streak >= 2:
                break
            r += 1
            continue
        blank_streak = 0
        for c, ccy in currencies:
            val = ws.cell(r, c).value
            if isinstance(val, (int, float)):
                # Some BMA sheets store decimals (0.045) while others store percent values (4.5).
                v = float(val)
                v = v * 100.0 if abs(v) <= 1.0 else v
                table[ccy][tenor] = round(v, 6)
            else:
                parsed = _as_float(val)
                if parsed is not None and abs(parsed) <= 1.0:
                    parsed *= 100.0
                table[ccy][tenor] = round(parsed, 6) if parsed is not None else None
        r += 1
    return table

def _bma_parse_discount_workbook(blob):
    if load_workbook is None:
        raise RuntimeError("openpyxl not available for BMA workbook parsing")

    wb = load_workbook(io.BytesIO(blob), data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]

    change_header = ""
    change_note = ""
    for r in range(1, 6):
        v = ws.cell(r, 2).value
        if isinstance(v, str) and "Changes for" in v:
            change_header = v.strip().rstrip(":")
            v2 = ws.cell(r + 1, 2).value
            if isinstance(v2, str):
                change_note = v2.strip()
            break

    risk_free = _bma_extract_table(ws, "Risk-Free Spot Rates")
    standard = _bma_extract_table(ws, "Standard Spot Rates")

    selected_tenors = ["0.5Y", "1Y", "2Y", "3Y", "5Y", "7Y", "10Y", "15Y", "20Y", "25Y", "30Y", "40Y", "50Y"]
    all_currencies = sorted(set(risk_free.keys()) | set(standard.keys()))

    currencies = {}
    for ccy in all_currencies:
        currencies[ccy] = {
            "risk_free_rates": [risk_free.get(ccy, {}).get(t) for t in selected_tenors],
            "standard_spot_rates": [standard.get(ccy, {}).get(t) for t in selected_tenors],
            # Preserve a simple legacy alias for downstream consumers.
            "rates": [standard.get(ccy, {}).get(t) for t in selected_tenors],
        }

    return {
        "sheet": ws.title,
        "change_header": change_header,
        "change_note": change_note,
        "tenors": selected_tenors,
        "currencies": currencies,
        "available_currencies": all_currencies,
    }

def _bma_load_quarter_url_cache():
    f = DATA / "bma_discount_url_cache.json"
    try:
        if f.exists():
            return json.loads(f.read_text())
    except Exception:
        pass
    return {}

def _bma_validate_quarter_payload(quarter):
    """Validate quarter payload and reject obvious absolute-value parse errors."""
    tenors = quarter.get("tenors", [])
    currencies = quarter.get("currencies", {})
    if not tenors or not currencies:
        return False, "missing tenors/currencies"

    non_null = 0
    for ccy, cdata in currencies.items():
        for key in ("risk_free_rates", "standard_spot_rates"):
            arr = cdata.get(key, [])
            if len(arr) != len(tenors):
                return False, f"{ccy} {key} length mismatch"
            for v in arr:
                if v is None:
                    continue
                non_null += 1
                # Guardrail for incorrectly-scaled values like 450 instead of 4.50.
                if v < -5 or v > 30:
                    return False, f"{ccy} {key} out-of-range value {v}"
    if non_null < 20:
        return False, "too few populated points"
    return True, "ok"

def _bma_save_quarter_url_cache(entries):
    f = DATA / "bma_discount_url_cache.json"
    cache = {}
    for e in entries:
        key = e["as_of_dt"].strftime("%Y-%m-%d")
        cache[key] = {
            "url": e.get("url", ""),
            "uploaded_on": e.get("uploaded_on", ""),
            "source_page": e.get("source_page", ""),
        }
    try:
        f.write_text(json.dumps(cache, indent=2))
    except Exception as e:
        log.warning(f"  BMA cache write failed: {e}")

def _bma_merge_cache_entries(entries):
    merged = {e["as_of_dt"].strftime("%Y-%m-%d"): e for e in entries}
    cache = _bma_load_quarter_url_cache()
    for k, v in cache.items():
        if k not in merged and v.get("url"):
            try:
                as_of_dt = datetime.strptime(k, "%Y-%m-%d")
            except Exception:
                continue
            uploaded_dt = parse_d(v.get("uploaded_on", "")) or _parse_iso_prefix_date_from_url(v.get("url", ""))
            merged[k] = {
                "as_of_dt": as_of_dt,
                "as_of": as_of_dt.strftime("%d %B %Y").lstrip("0"),
                "uploaded_dt": uploaded_dt,
                "uploaded_on": uploaded_dt.strftime("%d %B %Y").lstrip("0") if uploaded_dt else v.get("uploaded_on", ""),
                "url": v.get("url", ""),
                "source_page": v.get("source_page", "cache"),
            }
    out = list(merged.values())
    out.sort(key=lambda x: (x["as_of_dt"], x["uploaded_dt"] or datetime.min), reverse=True)
    return out

def _bma_bp_diff(a, b):
    if a is None or b is None:
        return None
    return round((a - b) * 100, 2)

def _bma_build_comparison(quarters):
    if not quarters:
        return {}
    latest = quarters[0]
    prev = quarters[1] if len(quarters) > 1 else None
    oldest = quarters[-1] if len(quarters) > 1 else None
    all_ccy = sorted(set().union(*[set(q.get("currencies", {}).keys()) for q in quarters]))
    tenors = latest.get("tenors", [])
    out = {}
    for ccy in all_ccy:
        rf_latest = latest.get("currencies", {}).get(ccy, {}).get("risk_free_rates", [None] * len(tenors))
        ss_latest = latest.get("currencies", {}).get(ccy, {}).get("standard_spot_rates", [None] * len(tenors))
        rf_prev = prev.get("currencies", {}).get(ccy, {}).get("risk_free_rates", [None] * len(tenors)) if prev else [None] * len(tenors)
        ss_prev = prev.get("currencies", {}).get(ccy, {}).get("standard_spot_rates", [None] * len(tenors)) if prev else [None] * len(tenors)
        rf_old = oldest.get("currencies", {}).get(ccy, {}).get("risk_free_rates", [None] * len(tenors)) if oldest else [None] * len(tenors)
        ss_old = oldest.get("currencies", {}).get(ccy, {}).get("standard_spot_rates", [None] * len(tenors)) if oldest else [None] * len(tenors)
        out[ccy] = {
            "risk_free_qoq_bp": [_bma_bp_diff(a, b) for a, b in zip(rf_latest, rf_prev)],
            "standard_spot_qoq_bp": [_bma_bp_diff(a, b) for a, b in zip(ss_latest, ss_prev)],
            "risk_free_vs_3q_ago_bp": [_bma_bp_diff(a, b) for a, b in zip(rf_latest, rf_old)],
            "standard_spot_vs_3q_ago_bp": [_bma_bp_diff(a, b) for a, b in zip(ss_latest, ss_old)],
        }
    return out


# ── 8. BMA RATES ──
def fetch_bma_rates():
    log.info("BMA RATES: fetching")
    manual_file = DATA / "bma_rates_manual.json"
    manual = json.loads(manual_file.read_text()) if manual_file.exists() else None

    entries = _bma_merge_cache_entries(_bma_discover_discount_files())
    _bma_save_quarter_url_cache(entries)

    latest = entries[0] if entries else None
    quarter_entries = []
    seen_quarters = set()
    for e in entries:
        qkey = e["as_of_dt"].strftime("%Y-%m-%d")
        if qkey not in seen_quarters:
            quarter_entries.append(e)
            seen_quarters.add(qkey)
        if len(quarter_entries) >= 4:
            break

    quarter_data = []
    for e in quarter_entries:
        parsed = None
        try:
            if e["url"].lower().endswith((".xlsx", ".xlsm", ".xls")):
                blob = get_bytes(e["url"], timeout=30, retries=2)
                parsed = _bma_parse_discount_workbook(blob)
                log.info(f"  BMA parsed workbook: {e['as_of']} ({e['url']})")
        except Exception as ex:
            log.warning(f"  BMA workbook parse failed for {e['as_of']}: {ex}")

        if parsed is None and manual and "quarters" in manual:
            parsed = manual["quarters"].get(e["as_of_dt"].strftime("%Y-%m-%d"))

        if parsed is not None:
            candidate = {
                "as_of_date": e["as_of_dt"].strftime("%Y-%m-%d"),
                "as_of_display": e["as_of"],
                "publication_date": e["uploaded_dt"].strftime("%Y-%m-%d") if e.get("uploaded_dt") else "",
                "publication_display": e.get("uploaded_on", ""),
                "quarter": _quarter_label(e["as_of_dt"]),
                "url": e["url"],
                "source_page": e.get("source_page", ""),
                "change_header": parsed.get("change_header", ""),
                "change_note": parsed.get("change_note", ""),
                "tenors": parsed.get("tenors", []),
                "available_currencies": parsed.get("available_currencies", []),
                "currencies": parsed.get("currencies", {}),
            }
            ok, reason = _bma_validate_quarter_payload(candidate)
            if not ok:
                log.warning(f"  BMA validation rejected {candidate['as_of_date']}: {reason}")
                continue
            quarter_data.append(candidate)
        else:
            quarter_data.append({
                "as_of_date": e["as_of_dt"].strftime("%Y-%m-%d"),
                "as_of_display": e["as_of"],
                "publication_date": e["uploaded_dt"].strftime("%Y-%m-%d") if e.get("uploaded_dt") else "",
                "publication_display": e.get("uploaded_on", ""),
                "quarter": _quarter_label(e["as_of_dt"]),
                "url": e["url"],
                "source_page": e.get("source_page", ""),
                "change_header": "",
                "change_note": "",
                "tenors": ["0.5Y", "1Y", "2Y", "3Y", "5Y", "7Y", "10Y", "15Y", "20Y", "25Y", "30Y", "40Y", "50Y"],
                "available_currencies": [],
                "currencies": {},
            })

    quarter_data = sorted(quarter_data, key=lambda q: q.get("as_of_date", ""), reverse=True)[:4]
    latest_q = quarter_data[0] if quarter_data else None
    comparison = _bma_build_comparison(quarter_data)

    # Preserve a compact top-level shape for downstream consumers.
    if latest_q:
        top_tenors = latest_q.get("tenors", ["0.5Y", "1Y", "2Y", "3Y", "5Y", "7Y", "10Y", "15Y", "20Y", "25Y", "30Y", "40Y", "50Y"])
        top_currencies = latest_q.get("currencies", {})
    else:
        top_tenors = ["0.5Y", "1Y", "2Y", "3Y", "5Y", "7Y", "10Y", "15Y", "20Y", "25Y", "30Y", "40Y", "50Y"]
        top_currencies = {}

    output = {
        "as_of_date": latest_q.get("as_of_display", "") if latest_q else (manual or {}).get("as_of_date", "Check BMA website"),
        "as_of_date_iso": latest_q.get("as_of_date", "") if latest_q else "",
        "publication_date": latest_q.get("publication_display", "") if latest_q else "",
        "publication_date_iso": latest_q.get("publication_date", "") if latest_q else "",
        "source": "BMA — EBS Discount Rates",
        "url": "https://www.bma.bm/documents-centre/documents-reporting-forms-and-guidelines/documents-insurance",
        "pdf_url": latest_q.get("url", "") if latest_q else "",
        "tenors": top_tenors,
        "currencies": top_currencies,
        "all_publications": [
            {
                "as_of": e["as_of"],
                "as_of_date": e["as_of_dt"].strftime("%Y-%m-%d"),
                "published": e.get("uploaded_on", ""),
                "published_date": e["uploaded_dt"].strftime("%Y-%m-%d") if e.get("uploaded_dt") else "",
                "url": e["url"],
                "source_page": e.get("source_page", ""),
            }
            for e in entries[:12]
        ],
        "quarter_history": quarter_data,
        "comparison": comparison,
        "note": (
            "Latest available BMA discount-rate workbook plus prior three quarter-end workbooks. "
            "Current top-level currencies are the latest quarter. "
            "For each currency, 'rates' aliases standard_spot_rates for backward compatibility."
        ),
    }

    # Optional manual override merge.
    if manual:
        if not output["currencies"] and manual.get("currencies"):
            output["currencies"] = manual["currencies"]
        if manual.get("tenors") and not latest_q:
            output["tenors"] = manual["tenors"]
        if manual.get("quarter_history") and not quarter_data:
            output["quarter_history"] = manual["quarter_history"]

    if not output["quarter_history"] and not output["currencies"]:
        last = load_last_bma()
        if last:
            output = last
            output["note"] = str(output.get("note", "")) + " Cache fallback used."

    write("bma_rates.json", output)
    log.info(f"  BMA RATES OK: {output.get('as_of_date_iso') or output.get('as_of_date','')}")

# ── 9. COMMODITIES ──
def fetch_commodities():
    log.info("COMMODITIES: fetching")
    TENOR_ORDER = ["3M", "6M", "12M", "24M"]
    TENOR_NS = [("3M", 3), ("6M", 6), ("12M", 12), ("24M", 24)]
    last = load_last_commodities()

    def get_spot(fred_series, inv_path, yahoo_front, te_url=None):
        obs = fred_csv(fred_series, start="2023-01-01")
        if obs:
            return round(obs[0]["value"], 2), obs[0]["date"], "FRED"
        if te_url:
            v = scrape_te_last_value(te_url)
            if v:
                return v, "", "TradingEconomics"
        v = scrape_commodity_spot(inv_path)
        if v:
            return v, "", "Investing"
        v = yahoo_price(yahoo_front)
        return v, "", "Yahoo"

    def fetch_all_futures(sym_fn, label_prefix):
        now_dt = datetime.utcnow()
        tasks = []
        for lbl, months in TENOR_NS:
            cur_sym, cur_exp = sym_fn(months, base_dt=now_dt)
            tasks.append((lbl, months, cur_sym, cur_exp, "now", now_dt))
            tasks.append((lbl, months, None, None, "1m", now_dt - timedelta(days=30)))
            tasks.append((lbl, months, None, None, "3m", now_dt - timedelta(days=91)))

        raw = {}

        def run(task):
            lbl, months, sym, exp, kind, target_dt = task
            if kind == "now":
                px = yahoo_price(sym)
                used_sym = sym
                if px is None and label_prefix == "USDINR":
                    px = yahoo_price("INR=F")
                    used_sym = "INR=F"
                return lbl, kind, px, "", used_sym, exp
            hist_sym, hist_exp = sym_fn(months, base_dt=target_dt)
            v, d = yahoo_price_near_date(hist_sym, target_dt)
            if v is None and label_prefix == "USDINR":
                v, d = yahoo_price_near_date("INR=F", target_dt)
                hist_sym = "INR=F"
            return lbl, kind, v, d, hist_sym, hist_exp

        with ThreadPoolExecutor(max_workers=12) as ex:
            fmap = {ex.submit(run, t): t for t in tasks}
            for fut in as_completed(fmap):
                try:
                    lbl, kind, price, d, sym, exp = fut.result()
                    raw.setdefault(lbl, {})[kind] = {"price": price, "date": d, "symbol": sym, "expiry": exp}
                    log.info(f"  {label_prefix} {lbl}/{kind}: {price} via {sym}")
                except Exception:
                    pass

        result = {}
        for lbl, months in TENOR_NS:
            r = raw.get(lbl, {})
            cur = r.get("now", {})
            p1m = r.get("1m", {})
            p3m = r.get("3m", {})
            result[lbl] = {
                "price": cur.get("price"),
                "expiry": cur.get("expiry"),
                "contract": (cur.get("symbol") or "").split(".")[0],
                "prior_1m": p1m.get("price"),
                "prior_1m_date": p1m.get("date"),
                "prior_1m_contract": (p1m.get("symbol") or "").split(".")[0],
                "prior_3m": p3m.get("price"),
                "prior_3m_date": p3m.get("date"),
                "prior_3m_contract": (p3m.get("symbol") or "").split(".")[0],
            }
        return {lbl: result[lbl] for lbl in TENOR_ORDER if lbl in result}

    gold_spot, gold_spot_date, gold_spot_source = get_spot("GOLDAMGBD228NLBM", "/commodities/gold", "GC=F", "https://tradingeconomics.com/commodity/gold")
    gold_1d, gold_1d_d = get_prior_spot("GOLDAMGBD228NLBM", "GC=F",   1, max_diff_days=5)
    gold_1m, gold_1m_d = get_prior_spot("GOLDAMGBD228NLBM", "GC=F",  30)
    gold_3m, gold_3m_d = get_prior_spot("GOLDAMGBD228NLBM", "GC=F",  91)
    gold_1y, gold_1y_d = get_prior_spot("GOLDAMGBD228NLBM", "GC=F", 365)
    gold_futures = fetch_all_futures(_gold_symbol, "Gold")

    wti_spot, wti_spot_date, wti_spot_source = get_spot("DCOILWTICO", "/commodities/crude-oil", "CL=F", "https://tradingeconomics.com/commodity/crude-oil")
    wti_1d, wti_1d_d   = get_prior_spot("DCOILWTICO",   "CL=F",   1, max_diff_days=5)
    wti_1m, wti_1m_d   = get_prior_spot("DCOILWTICO",   "CL=F",  30)
    wti_3m, wti_3m_d   = get_prior_spot("DCOILWTICO",   "CL=F",  91)
    wti_1y, wti_1y_d   = get_prior_spot("DCOILWTICO",   "CL=F", 365)
    wti_futures = fetch_all_futures(_wti_symbol, "WTI")

    brent_spot, brent_spot_date, brent_spot_source = get_spot("DCOILBRENTEU", "/commodities/brent-oil", "BZ=F", "https://tradingeconomics.com/commodity/brent-crude-oil")
    brent_1d, brent_1d_d = get_prior_spot("DCOILBRENTEU", "BZ=F",   1, max_diff_days=5)
    brent_1m, brent_1m_d = get_prior_spot("DCOILBRENTEU", "BZ=F",  30)
    brent_3m, brent_3m_d = get_prior_spot("DCOILBRENTEU", "BZ=F",  91)
    brent_1y, brent_1y_d = get_prior_spot("DCOILBRENTEU", "BZ=F", 365)
    brent_futures = fetch_all_futures(_brent_symbol, "Brent")

    if last:
        if gold_spot is None:
            gold_spot = last.get("gold", {}).get("spot")
        if wti_spot is None:
            wti_spot = last.get("wti", {}).get("spot")
        if brent_spot is None:
            brent_spot = last.get("brent", {}).get("spot")

    # USD/INR spot and history from FRED DEXINUS (Indian Rupees per 1 USD)
    # Fallback for latest spot uses manual market scrape if FRED is stale/unavailable.
    usdinr_obs = fred_csv("DEXINUS", start="2023-01-01")
    if not usdinr_obs and last:
        usdinr_obs = []
    if usdinr_obs:
        usdinr_spot = round(usdinr_obs[0]["value"], 4)
        usdinr_spot_date = usdinr_obs[0]["date"]
        usdinr_spot_source = "FRED DEXINUS"
        def _usdinr_prior(days_back, max_diff=14):
            target = datetime.utcnow() - timedelta(days=days_back)
            best = min(usdinr_obs, key=lambda o: abs((datetime.strptime(o["date"], "%Y-%m-%d") - target).days))
            diff = abs((datetime.strptime(best["date"], "%Y-%m-%d") - target).days)
            return (round(best["value"], 4), best["date"]) if diff <= max_diff else (None, None)
        usdinr_1d, usdinr_1d_d = _usdinr_prior(1, max_diff=5)
        usdinr_1m, usdinr_1m_d = _usdinr_prior(30)
        usdinr_3m, usdinr_3m_d = _usdinr_prior(91)
        usdinr_1y, usdinr_1y_d = _usdinr_prior(365)
    else:
        usdinr_spot = last.get("usdinr", {}).get("spot") if last else None
        usdinr_spot_date = last.get("usdinr", {}).get("spot_date", "") if last else ""
        usdinr_spot_source = "cache"
        usdinr_1d = usdinr_1d_d = usdinr_1m = usdinr_1m_d = None
        usdinr_3m = usdinr_3m_d = usdinr_1y = usdinr_1y_d = None

    # If FRED spot is stale (>2 days old) or missing, use manual scrape for latest spot.
    try:
        spot_dt = datetime.strptime(usdinr_spot_date, "%Y-%m-%d") if usdinr_spot_date else None
        spot_age_days = (datetime.utcnow() - spot_dt).days if spot_dt else 999
    except Exception:
        spot_age_days = 999
    if usdinr_spot is None or spot_age_days > 2:
        scraped_usdinr = scrape_fx_spot("/currencies/usd-inr")
        if scraped_usdinr is not None:
            usdinr_spot = scraped_usdinr
            usdinr_spot_date = datetime.utcnow().strftime("%Y-%m-%d")
            usdinr_spot_source = "Investing scrape"

    usdinr_futures = fetch_all_futures(_usdinr_symbol, "USDINR")
    if not any((usdinr_futures.get(t, {}) or {}).get("price") is not None for t in ["3M", "6M", "12M", "24M"]):
        fwds = scrape_usdinr_forwards(usdinr_spot)
        if fwds:
            for t in ["3M", "6M", "12M", "24M"]:
                if t not in usdinr_futures:
                    usdinr_futures[t] = {}
                if usdinr_futures[t].get("price") is None and fwds.get(t) is not None:
                    usdinr_futures[t]["price"] = fwds[t]
                    usdinr_futures[t]["contract"] = "INV-FWD"
                    usdinr_futures[t]["expiry"] = t

    write("commodities.json", {
        "date": datetime.utcnow().strftime("%Y-%m-%d"),
        "source": "FRED / TradingEconomics / Investing / Yahoo Finance",
        "gold": {
            "spot": gold_spot,
            "spot_date": gold_spot_date,
            "spot_source": gold_spot_source,
            "unit": "USD/troy oz",
            "prior_1d": gold_1d, "prior_1d_date": gold_1d_d,
            "prior_1m": gold_1m, "prior_1m_date": gold_1m_d,
            "prior_3m": gold_3m, "prior_3m_date": gold_3m_d,
            "prior_1y": gold_1y, "prior_1y_date": gold_1y_d,
            "futures": gold_futures
        },
        "wti": {
            "spot": wti_spot,
            "spot_date": wti_spot_date,
            "spot_source": wti_spot_source,
            "unit": "USD/barrel",
            "prior_1d": wti_1d,   "prior_1d_date": wti_1d_d,
            "prior_1m": wti_1m,   "prior_1m_date": wti_1m_d,
            "prior_3m": wti_3m,   "prior_3m_date": wti_3m_d,
            "prior_1y": wti_1y,   "prior_1y_date": wti_1y_d,
            "futures": wti_futures
        },
        "brent": {
            "spot": brent_spot,
            "spot_date": brent_spot_date,
            "spot_source": brent_spot_source,
            "unit": "USD/barrel",
            "prior_1d": brent_1d, "prior_1d_date": brent_1d_d,
            "prior_1m": brent_1m, "prior_1m_date": brent_1m_d,
            "prior_3m": brent_3m, "prior_3m_date": brent_3m_d,
            "prior_1y": brent_1y, "prior_1y_date": brent_1y_d,
            "futures": brent_futures
        },
        "usdinr": {
            "spot": usdinr_spot,
            "spot_date": usdinr_spot_date,
            "spot_source": usdinr_spot_source,
            "unit": "INR per USD",
            "prior_1d": usdinr_1d, "prior_1d_date": usdinr_1d_d,
            "prior_1m": usdinr_1m, "prior_1m_date": usdinr_1m_d,
            "prior_3m": usdinr_3m, "prior_3m_date": usdinr_3m_d,
            "prior_1y": usdinr_1y, "prior_1y_date": usdinr_1y_d,
            "futures": usdinr_futures,
        },
        "note": "Spot history comes from daily FRED series. Futures history is roll-aware by target date. USD/INR uses FRED DEXINUS with Investing scrape fallback for latest spot."
    })
    log.info("  COMMODITIES OK")

# ── RUN ──
def main():
    log.info("=" * 50)
    results = {}
    for name, fn in [
        ("ust", fetch_ust),
        ("jgb", fetch_jgb),
        ("gilt", fetch_gilt),
        ("eur", fetch_eur),
        ("india", fetch_india),
        ("credit", fetch_credit),
        ("sofr", fetch_sofr),
        ("bma_rates", fetch_bma_rates),
        ("commodities", fetch_commodities),
    ]:
        try:
            fn()
            results[name] = "ok"
        except Exception as e:
            log.error(f"  {name} FAILED: {e}")
            results[name] = str(e)
    write("manifest.json", {"results": results, "run": datetime.utcnow().isoformat() + "Z"})
    failed = [k for k, v in results.items() if v != "ok"]
    log.info(f"Done: {len(results)-len(failed)}/{len(results)} ok" + (f", failed: {failed}" if failed else ""))

if __name__ == "__main__":
    main()
